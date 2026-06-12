from pathlib import Path
import json
from collections import defaultdict
import pandas as pd
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# ============================================================
# HELPERS
# ============================================================


def safe_table_name(name: str, idx: int) -> str:
    clean = re.sub(r"[^A-Za-z0-9_]", "_", name)
    clean = clean.strip("_")[:20]
    return f"T_{clean}_{idx}"


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 4


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="DDDDDD")


# ============================================================
# MAIN FUNCTION
# ============================================================


def create_clash_tables(input_dir, excel_file="clash_summary.xlsx"):
    input_dir = Path(input_dir)

    # --------------------------------------------------------
    # STORAGE
    # --------------------------------------------------------
    type_counts = defaultdict(int)
    file_counts = defaultdict(int)

    # file -> element -> type -> count
    element_counts = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    all_types = set()
    all_files = set()

    # --------------------------------------------------------
    # READ JSON FILES
    # --------------------------------------------------------
    for json_file in input_dir.glob("*.json"):
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not data:
            continue

        report = data[0]

        file_a = Path(report["a"][0]["file"]).stem
        file_b = Path(report["b"][0]["file"]).stem

        all_files.update([file_a, file_b])

        clashes = report.get("clashes", {})

        pair_files = tuple(sorted([file_a, file_b]))
        file_counts[pair_files] += len(clashes)

        for clash in clashes.values():
            a_id = clash.get("a_global_id")
            b_id = clash.get("b_global_id")

            a_type = clash.get("a_ifc_class", "Unknown")
            b_type = clash.get("b_ifc_class", "Unknown")

            a_name = clash.get("a_name", "<No Name>")
            b_name = clash.get("b_name", "<No Name>")

            all_types.update([a_type, b_type])

            # TYPE MATRIX
            pair_types = tuple(sorted([a_type, b_type]))
            type_counts[pair_types] += 1

            # ELEMENT STATS (Option A structure)
            if a_id:
                element_counts[file_a][a_id]["name"] = a_name
                element_counts[file_a][a_id]["type"] = a_type
                element_counts[file_a][a_id][b_type] += 1

            if b_id:
                element_counts[file_b][b_id]["name"] = b_name
                element_counts[file_b][b_id]["type"] = b_type
                element_counts[file_b][b_id][a_type] += 1

    # --------------------------------------------------------
    # TYPE MATRIX
    # --------------------------------------------------------
    types = sorted(all_types)
    type_df = pd.DataFrame(0, index=types, columns=types, dtype=int)

    for (t1, t2), count in type_counts.items():
        if t1 == t2:
            type_df.loc[t1, t2] += count
        else:
            type_df.loc[t1, t2] += count
            type_df.loc[t2, t1] += count

    # --------------------------------------------------------
    # FILE MATRIX
    # --------------------------------------------------------
    files = sorted(all_files)
    file_df = pd.DataFrame(0, index=files, columns=files, dtype=int)

    for (f1, f2), count in file_counts.items():
        if f1 == f2:
            file_df.loc[f1, f2] += count
        else:
            file_df.loc[f1, f2] += count
            file_df.loc[f2, f1] += count

    # --------------------------------------------------------
    # WRITE EXCEL
    # --------------------------------------------------------
    output_path = input_dir / excel_file

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        type_df.to_excel(writer, sheet_name="IFC_Type_Matrix")
        file_df.to_excel(writer, sheet_name="File_Matrix")

        # dynamic per-file sheets
        for file_name in files:
            rows = []

            for eid, data in element_counts[file_name].items():
                element_type = data.get("type", "Unknown")
                element_name = data.get("name", "<No Name>")

                clash_count = sum(
                    v for k, v in data.items() if k not in ["name", "type"]
                )

                row = {
                    "ElementID": eid,
                    "ElementName": element_name,
                    "ElementType": element_type,
                    "ClashCount": clash_count,
                }

                for k, v in data.items():
                    if k not in ["name", "type"]:
                        row[k] = v

                rows.append(row)

            df = pd.DataFrame(rows).fillna(0)

            sheet = file_name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)

    # --------------------------------------------------------
    # POST PROCESS
    # --------------------------------------------------------
    wb = load_workbook(output_path)

    green = "C6EFCE"
    yellow = "FFEB84"
    red = "F8696B"

    # --------------------------------------------------------
    # TYPE MATRIX HEATMAP
    # --------------------------------------------------------
    ws = wb["IFC_Type_Matrix"]

    ws.conditional_formatting.add(
        f"B2:{get_column_letter(len(types) + 1)}{len(types) + 1}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color=green,
            mid_type="percentile",
            mid_value=50,
            mid_color=yellow,
            end_type="max",
            end_color=red,
        ),
    )

    style_header(ws)
    autosize(ws)

    # --------------------------------------------------------
    # FILE MATRIX HEATMAP
    # --------------------------------------------------------
    ws = wb["File_Matrix"]

    ws.conditional_formatting.add(
        f"B2:{get_column_letter(len(files) + 1)}{len(files) + 1}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color=green,
            mid_type="percentile",
            mid_value=50,
            mid_color=yellow,
            end_type="max",
            end_color=red,
        ),
    )

    style_header(ws)
    autosize(ws)

    # --------------------------------------------------------
    # FILE SHEETS FORMATTING
    # --------------------------------------------------------
    table_idx = 0

    for file_name in files:
        ws = wb[file_name[:31]]

        if ws.max_row < 2:
            continue

        table_idx += 1

        table = Table(
            displayName=safe_table_name(file_name, table_idx),
            ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
        )

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=False,
        )

        ws.add_table(table)

        ws.freeze_panes = "A2"
        style_header(ws)
        autosize(ws)

    # --------------------------------------------------------
    # DASHBOARD
    # --------------------------------------------------------
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]

    ws = wb.create_sheet("Dashboard", 0)

    total_clashes = sum(type_counts.values())
    worst_pair = max(file_counts, key=file_counts.get, default=None)

    worst_pair_str = (
        f"{worst_pair[0]} ↔ {worst_pair[1]} ({file_counts[worst_pair]})"
        if worst_pair
        else "N/A"
    )

    ws.append(["BIM CLASH DASHBOARD"])
    ws.append([])

    ws.append(["Total Clashes", total_clashes])
    ws.append(["Files", len(files)])
    ws.append(["IFC Types", len(types)])
    ws.append(["Worst File Pair", worst_pair_str])

    ws.append([])
    ws.append(["TOP CLASHING ELEMENTS"])

    ws.append(["File", "ElementID", "ElementName", "ElementType", "ClashCount"])

    # flatten
    flat = []

    for f, elems in element_counts.items():
        for eid, data in elems.items():
            clash_count = sum(v for k, v in data.items() if k not in ["name", "type"])

            flat.append(
                [
                    f,
                    eid,
                    data.get("name", "<No Name>"),
                    data.get("type", "Unknown"),
                    clash_count,
                ]
            )

    top = sorted(flat, key=lambda x: x[4], reverse=True)[:25]

    for row in top:
        ws.append(row)

    style_header(ws)
    autosize(ws)

    wb.save(output_path)

    print(f"Excel report created: {output_path}")
